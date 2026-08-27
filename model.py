from dataclasses import dataclass, field
from typing import Optional
import openpyxl


@dataclass
class OSSEntry:
    row_num: int          # 最初に登場した行番号
    oss_name: str
    license_names: list[str]   # E列: 複数ライセンス対応
    copyrights: list[str]      # AA列: 複数著作権者対応
    license_texts: list[str]   # AB列: 複数行ライセンス原文対応


@dataclass
class OSSData:
    entries: list[OSSEntry] = field(default_factory=list)

    def add(self, entry: OSSEntry):
        self.entries.append(entry)

    def last(self) -> Optional[OSSEntry]:
        return self.entries[-1] if self.entries else None


# 列番号定数（1始まり）
COL_A  = 1   # 終了判定用
COL_B  = 2   # OSS名
COL_E  = 5   # ライセンス名
COL_AA = 27  # Copyright
COL_AB = 28  # ライセンス原文

# 空欄チェック対象（B列のみ必須、他は追記判定後にチェック）
REQUIRED_COL = {COL_B: "B"}

START_ROW = 9


# ------------------------------------------------------------------ #
# ダミーワークシート                                                    #
# ------------------------------------------------------------------ #

# ダミーデータ定義
# 各行は dict[col_num -> value]
# B列が同じ → 追記ケースをテスト
_DUMMY_ROWS = [
    # ── requests: 通常ケース（1行完結）
    {COL_A: 1, COL_B: "requests",
     COL_E: "Apache-2.0",
     COL_AA: "Copyright 2023 Kenneth Reitz",
     COL_AB: "Apache License\nVersion 2.0, January 2004\nFull license text here."},

    # ── numpy: 準正常系ケース1 著作権者が複数行にまたがる
    {COL_A: 2, COL_B: "numpy",
     COL_E: "BSD-3-Clause",
     COL_AA: "Copyright (c) NumPy Developers",
     COL_AB: "BSD 3-Clause License\n\nRedistribution and use in source..."},
    {COL_A: 3, COL_B: "numpy",   # 同じOSS名 → AA列を追記
     COL_E: None,                # E列なし → ライセンス名は追記しない
     COL_AA: "Copyright 2010 Pallets",
     COL_AB: None},              # AB列なし → ライセンス原文は追記しない

    # ── biglib: 準正常系ケース2 ライセンス原文が複数行にまたがる
    {COL_A: 4, COL_B: "biglib",
     COL_E: "GPL-3.0",
     COL_AA: "Copyright (C) 2007 Free Software Foundation",
     COL_AB: "GNU GENERAL PUBLIC LICENSE\nVersion 3, 29 June 2007\n\n"
             "Everyone is permitted to copy and distribute verbatim copies\n"
             "of this license document, but changing it is not allowed.\n\n"
             "a copy of the Program in return for a fee."},
    {COL_A: 5, COL_B: "biglib",  # 同じOSS名 → AB列を前行に続けて追記
     COL_E: None,
     COL_AA: None,
     COL_AB: "END OF TERMS AND CONDITIONS"},

    # ── multilic: 準正常系ケース3 複数ライセンスをもつOSS
    {COL_A: 6, COL_B: "multilic",
     COL_E: "MIT",
     COL_AA: "Copyright (c) 2020 Example Corp",
     COL_AB: "MIT License\n\nPermission is hereby granted..."},
    {COL_A: 7, COL_B: "multilic",  # 同じOSS名 → E列をカンマ区切りで追記
     COL_E: "BSD-3-Clause",
     COL_AA: None,
     COL_AB: None},
    {COL_A: 8, COL_B: "multilic",
     COL_E: "GPL-2.0",
     COL_AA: None,
     COL_AB: None},

    # ── dirtylib: 準正常系ケース5 _x000C_ が混入しているケース
    {COL_A: 9, COL_B: "dirtylib_x000C_",
     COL_E: "MIT_x000C_",
     COL_AA: "Copyright_x000C_ 2024 Dirty Corp",
     COL_AB: "MIT License_x000C_\n\nPermission is hereby granted_x000C_..."},
    {COL_A: 10, COL_B: "dirtylib_x000C_",  # 追記行にも混入
     COL_E: None,
     COL_AA: "Copyright_x000C_ 2025 Another Author",
     COL_AB: "Additional terms_x000C_ here."},

    # ── complexlib: 準正常系ケース6 準正常系1～3の全部の組み合わせ
    #   著作権者が2行（準正常系ケース1）
    #   ライセンス原文が3行に分割（準正常系ケース2）
    #   ライセンス名が2種類（準正常系ケース3）
    {COL_A:  11, COL_B: "complexlib",
     COL_E:  "LGPL-2.1",
     COL_AA: "Copyright (C) 1991 Free Software Foundation",
     COL_AB: "GNU LESSER GENERAL PUBLIC LICENSE\nVersion 2.1, February 1999\n\n"
             "Everyone is permitted to copy and distribute verbatim copies [part1]"},
    {COL_A: 12, COL_B: "complexlib",  # 準正常系ケース1: 著作権者追加 / 準正常系ケース2: 原文続き / 準正常系ケース3: ライセンス追加
     COL_E:  "MIT",
     COL_AA: "Copyright (c) 2001 Example Contributor",
     COL_AB: "of this license document, but changing it is not allowed. [part2]"},
    {COL_A: 13, COL_B: "complexlib",  # 準正常系ケース2: 原文続き（E列・AA列はなし）
     COL_E:  None,
     COL_AA: None,
     COL_AB: "END OF TERMS AND CONDITIONS [part3]"},

    # ── scipy: 集約ケース1 numpyと同じBSD-3-Clause原文 → 集約で同一ブロックになる
    #   numpyと著作権者は異なるが、ライセンス原文が完全一致するため集約される
    {COL_A: 14, COL_B: "scipy",
     COL_E:  "BSD-3-Clause",
     COL_AA: "Copyright (c) 2001-2002 Enthought, Inc. 2003-2024, SciPy Developers",
     COL_AB: "BSD 3-Clause License\n\nRedistribution and use in source..."},

    # ── 集約フォーマット用テストデータ
    # requests-cache: requestsと同じApache-2.0原文 → 集約で同一ブロックになる
    {COL_A: 15, COL_B: "requests-cache",
     COL_E:  "Apache-2.0",
     COL_AA: "Copyright 2021 Roman Haritonov",
     COL_AB: "Apache License\nVersion 2.0, January 2004\nFull license text here."},

    # flask: multilicと同じMIT原文 → 集約で同一ブロックになる
    {COL_A: 16, COL_B: "flask",
     COL_E:  "MIT",
     COL_AA: "Copyright 2010 Pallets",
     COL_AB: "MIT License\n\nPermission is hereby granted..."},
]


class _Cell:
    def __init__(self, value):
        self.value = value


class DummyWorksheet:
    def __init__(self):
        self._data: dict[tuple[int, int], object] = {}
        for i, row_dict in enumerate(_DUMMY_ROWS):
            r = START_ROW + i
            for col, val in row_dict.items():
                self._data[(r, col)] = val
        self.max_row = START_ROW + len(_DUMMY_ROWS) - 1

    def cell(self, row: int, column: int) -> _Cell:
        return _Cell(self._data.get((row, column), None))


# ------------------------------------------------------------------ #
# Excel / ダミー読み込み                                               #
# ------------------------------------------------------------------ #

def _sanitize(value) -> str:
    """セル値を文字列化し、_x000C_（フォームフィード）を除去して返す。前後スペースは保持。"""
    return str(value).replace("_x000C_", "")


def _is_blank(value) -> bool:
    """セル値が空欄（None または空白のみ）かどうかを返す。"""
    return value is None or _sanitize(value).strip() == ""


def load_excel(filepath: str, sheetname: str, use_dummy: bool = False,
               start_row: int = START_ROW, end_row: Optional[int] = None) -> OSSData:
    if use_dummy:
        ws = DummyWorksheet()
    else:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        if sheetname not in wb.sheetnames:
            raise ValueError(
                f"シート '{sheetname}' が見つかりません。利用可能なシート: {wb.sheetnames}"
            )
        ws = wb[sheetname]

    return _parse_worksheet(ws, start_row=start_row, end_row=end_row)


def _parse_worksheet(ws, start_row: int = START_ROW,
                     end_row: Optional[int] = None) -> OSSData:
    data = OSSData()
    last_row = end_row if end_row is not None else ws.max_row

    for row_idx in range(start_row, last_row + 1):
        # end_row 未指定のときはA列空欄で終端判定
        if end_row is None:
            a_val = ws.cell(row=row_idx, column=COL_A).value
            if a_val is None or str(a_val).strip() == "":
                break

        # B列（OSS名）は必須
        b_val = ws.cell(row=row_idx, column=COL_B).value
        if b_val is None or str(b_val).strip() == "":
            raise ValueError(
                f"{row_idx}行, B列が空欄であるため中断、空欄は無いようにしてください"
            )
        oss_name = _sanitize(b_val)

        last = data.last()

        if last is not None and last.oss_name == oss_name:
            # ── 同じOSS名 → 追記処理
            _append_row(ws, row_idx, last)
        else:
            # ── 新しいOSS → 新規エントリ作成（E/AA/AB列は必須チェック）
            entry = _new_entry(ws, row_idx, oss_name)
            data.add(entry)

    return data


def _new_entry(ws, row_idx: int, oss_name: str) -> OSSEntry:
    """新規エントリ作成。E/AA/AB列は必須。"""
    cols = {
        COL_E:  "E",
        COL_AA: "AA",
        COL_AB: "AB",
    }
    values = {}
    for col, col_name in cols.items():
        val = ws.cell(row=row_idx, column=col).value
        if _is_blank(val):
            raise ValueError(
                f"{row_idx}行, {col_name}列が空欄であるため中断、空欄は無いようにしてください"
            )
        values[col] = _sanitize(val)

    return OSSEntry(
        row_num=row_idx,
        oss_name=oss_name,
        license_names=[values[COL_E]],
        copyrights=[values[COL_AA]],
        license_texts=[values[COL_AB]],
    )


def _append_row(ws, row_idx: int, entry: OSSEntry):
    """同一OSS名の追加行を既存エントリに追記する。"""
    e_val  = ws.cell(row=row_idx, column=COL_E).value
    aa_val = ws.cell(row=row_idx, column=COL_AA).value
    ab_val = ws.cell(row=row_idx, column=COL_AB).value

    # E列: 記載があればカンマ区切りで追加
    if not _is_blank(e_val):
        entry.license_names.append(_sanitize(e_val))

    # AA列: 記載があれば改行区切りで追加
    if not _is_blank(aa_val):
        entry.copyrights.append(_sanitize(aa_val))

    # AB列: 記載があれば続けて追記（改行で連結）
    if not _is_blank(ab_val):
        entry.license_texts.append(_sanitize(ab_val))
